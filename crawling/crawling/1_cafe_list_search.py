# -*- coding: utf-8 -*-
"""
네이버 지도 placeId 크롤러 (강릉 카페 검색 전용)

- "강릉 카페" 검색 후, 결과 목록 클릭 → entryIframe에서 이름/주소 추출
- 저장 스키마: [no, store_name, store_url_naver]
"""

import os, re, time, urllib.parse
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ===== 사용자 설정 =====
OUTPUT_PATH = "naver_cafe_list.xlsx"
CITY_FILTER = "강릉"
SEARCH_KEYWORD = "강릉 카페"


# ===== 공통 유틸 =====
def build_review_url(place_id: str) -> str:
    return f"https://m.place.naver.com/restaurant/{place_id}/review/visitor?entry=ple&reviewSort=recent"


def extract_place_id_from_url(url: str) -> str | None:
    m = re.search(r"/place/(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"[?&#]id=(\d+)", url)
    return m.group(1) if m else None


# ===== 저장 =====
def append_single_row(path: str, no: int, store_name: str, store_url_naver: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["no", "store_name", "store_url_naver"])
        wb.save(path)
        print(f"📄 새 파일 생성 + 헤더 기록: {path}")

    wb = load_workbook(path)
    ws = wb.active
    ws.append([no, store_name, store_url_naver])
    wb.save(path)
    print(f"📝 저장완료 | no={no}, store_name='{store_name}'")


# ===== 프레임 =====
def _switch_to_search_iframe(driver, wait):
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
    print("✅ searchIframe 전환 성공")


def _switch_to_entry_iframe(driver, wait):
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "entryIframe")))


# ===== 목록 =====
def _scroll_all_in_list(driver):
    try:
        container = driver.find_element(
            By.CSS_SELECTOR, "div#_pcmap_list_scroll_container"
        )
    except:
        return
    prev_h = 0
    stall = 0
    while True:
        driver.execute_script(
            "arguments[0].scrollTop = arguments[0].scrollHeight;", container
        )
        time.sleep(0.5)
        h = driver.execute_script("return arguments[0].scrollHeight;", container)
        stall = stall + 1 if h == prev_h else 0
        if stall >= 3:
            break
        prev_h = h


def _find_list_items(driver):
    li_selectors = [
        "div#_pcmap_list_scroll_container ul > li",
        "ul > li.VYGLG",  # 최신 구조
        "ul > li.UEzoS",  # 이전 구조
        "ul > li",
    ]
    for sel in li_selectors:
        lis = driver.find_elements(By.CSS_SELECTOR, sel)
        if lis:
            print(f"✅ 셀렉터 '{sel}' 로 {len(lis)}개 항목 탐지")
            return lis
    return []


# ===== entryIframe 정보 추출 =====
def _extract_name_from_entry_iframe(driver) -> str:
    for sel in ["#_title > div > span.GHAhO", "span.TYaxT", "span.Fc1rA", "h2 span"]:
        try:
            text = driver.find_element(By.CSS_SELECTOR, sel).text.strip()
            if text:
                return text
        except:
            continue
    return ""


def _extract_address_from_entry_iframe(driver) -> str:
    for sel in [
        "span.LDgIH",
        "span.Fc1rA",
        "div.O8qbU span",
        "div.place_section_content span",
        "div.detail_address span",
        'span[class*="addr"]',
        'div[class*="address"]',
    ]:
        try:
            for elem in driver.find_elements(By.CSS_SELECTOR, sel):
                text = elem.text.strip()
                if text and len(text) > 5:
                    return text
        except:
            continue
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        for line in body_text.split("\n"):
            if ("강원" in line or "강릉" in line) and len(line) > 10:
                return line.strip()
    except:
        pass
    return ""


# ===== 클릭 대상 =====
def _find_clickable_link(li):
    selectors = [
        "a.place_bluelink",
        "a.tit_name",
        "a[href*='/place/']",
        "a",
        "div[role='button']",
    ]
    for sel in selectors:
        try:
            return li.find_element(By.CSS_SELECTOR, sel)
        except:
            continue
    return None


# ===== 페이지네이션 =====
def _get_current_page_number(driver):
    try:
        current_btn = driver.find_element(
            By.CSS_SELECTOR, 'a.mBN2s.qxokY, a.mBN2s[aria-current="true"]'
        )
        return int(current_btn.text.strip())
    except:
        return 1


def _has_more_pages(driver) -> bool:
    # 다음 숫자 버튼 있으면 True
    try:
        current_page = _get_current_page_number(driver)
        buttons = driver.find_elements(By.CSS_SELECTOR, "a.mBN2s")
        for btn in buttons:
            try:
                if int(btn.text.strip()) > current_page:
                    return True
            except:
                continue
    except:
        pass

    # "다음페이지" 버튼이 활성화 상태면 True
    try:
        next_arrow = driver.find_element(By.CSS_SELECTOR, "a.eUTY2")
        return next_arrow.get_attribute("aria-disabled") == "false"
    except:
        return False


def _click_next_page(driver, wait):
    current_page = _get_current_page_number(driver)
    next_page = current_page + 1

    # 1) 숫자 버튼 클릭
    try:
        buttons = driver.find_elements(By.CSS_SELECTOR, "a.mBN2s")
        for btn in buttons:
            if btn.text.strip() == str(next_page):
                print(f"   🔘 {next_page}페이지 버튼 클릭")
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", btn
                )
                time.sleep(0.3)
                btn.click()
                return True
    except:
        pass

    # 2) '다음페이지' 화살표 클릭
    try:
        next_arrow = driver.find_element(
            By.CSS_SELECTOR, 'a.eUTY2[aria-disabled="false"]'
        )
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", next_arrow
        )
        time.sleep(0.3)
        next_arrow.click()
        print("➡️ '다음페이지' 버튼 클릭")
        return True
    except:
        pass

    return False


# ===== 메인 =====
def main():
    print("🚀 시작")

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 15)

    no = 1
    total_saved = total_skipped = 0

    keyword = SEARCH_KEYWORD
    print(f"\n🔎 검색 시작: {keyword}")

    encoded = urllib.parse.quote(keyword, safe="")
    url = f"https://map.naver.com/p/search/{encoded}"
    driver.get(url)
    time.sleep(2.0)

    try:
        _switch_to_search_iframe(driver, wait)
    except TimeoutException:
        print("⚠️ searchIframe 진입 실패")
        driver.quit()
        return

    page_num = 1

    while True:
        print(f"\n📄 페이지 {page_num} 처리 중...")

        try:
            wait.until(
                EC.presence_of_element_located(
                    (
                        By.CSS_SELECTOR,
                        "div#_pcmap_list_scroll_container ul > li, ul > li.VYGLG, ul > li.UEzoS, ul > li",
                    )
                )
            )
        except TimeoutException:
            print("⚠️ 리스트 로드 실패 (타임아웃)")
            break

        _scroll_all_in_list(driver)
        lis = _find_list_items(driver)
        if not lis:
            print("⚠️ 목록(li) 탐색 실패")
            break

        saved = skipped = 0

        for idx, li in enumerate(lis, start=1):
            try:
                click_target = _find_clickable_link(li)
                if click_target is None:
                    print(f"  • #{idx} 클릭 요소 없음 → 스킵")
                    skipped += 1
                    continue

                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", li
                )
                time.sleep(0.3)
                click_target.click()
                time.sleep(0.5)

                _switch_to_entry_iframe(driver, wait)
                WebDriverWait(driver, 8).until(
                    lambda d: re.search(r"/place/\d+", d.current_url)
                )

                current = driver.current_url
                pid = extract_place_id_from_url(current)
                if not pid:
                    _switch_to_search_iframe(driver, wait)
                    skipped += 1
                    continue

                store_name = _extract_name_from_entry_iframe(driver)
                addr = _extract_address_from_entry_iframe(driver)

                if CITY_FILTER not in addr:
                    print(f"  • #{idx} {store_name} → 주소 미일치 스킵 | addr='{addr}'")
                    _switch_to_search_iframe(driver, wait)
                    skipped += 1
                    continue

                review_url = build_review_url(pid)
                append_single_row(OUTPUT_PATH, no, store_name, review_url)
                print(f"  ✅ 저장 • no={no} | {store_name} [{addr}] → {pid}")
                no += 1
                saved += 1

                _switch_to_search_iframe(driver, wait)
                time.sleep(0.3)

            except Exception as e:
                print(f"  • #{idx} 오류: {e} → 계속")
                skipped += 1
                try:
                    _switch_to_search_iframe(driver, wait)
                except:
                    pass
                continue

        total_saved += saved
        total_skipped += skipped
        print(f"📊 페이지 {page_num} 완료: 저장 {saved}건, 스킵 {skipped}건")

        if not _has_more_pages(driver):
            print("✋ 마지막 페이지 도달!")
            break

        try:
            if _click_next_page(driver, wait):
                try:
                    wait.until(
                        EC.presence_of_element_located(
                            (
                                By.CSS_SELECTOR,
                                "div#_pcmap_list_scroll_container ul > li, ul > li.VYGLG, ul > li.UEzoS, ul > li",
                            )
                        )
                    )
                except TimeoutException:
                    print("⚠️ 다음 페이지 로드 실패")
                    break
                time.sleep(1.5)
                page_num += 1
            else:
                print("⚠️ 다음 페이지 버튼 클릭 실패")
                break
        except Exception as e:
            print(f"⚠️ 다음 페이지 이동 실패: {e}")
            break

    driver.quit()
    print(f"\n🎉 전체 크롤링 완료!")
    print(f"📊 총 {page_num}개 페이지 처리")
    print(f"✅ 저장: {total_saved}건")
    print(f"⏭️ 스킵: {total_skipped}건")
    print(f"📄 결과 파일: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
